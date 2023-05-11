import datetime
import os
import sys
import pandas as pd
import openpyxl as opx
import win32com.client as win32
import pythoncom
import logging
import config


def get_project(path):
    """Открывает файл проекта и возвращает объект проекта"""
    if not os.path.isabs(path):
        logging.warning('%s: Путь до файла проекта не абсолютный', get_project.__name__)
    logging.info('%s: Пытаемся открыть файл проекта', get_project.__name__)
    try:
        msp = win32.Dispatch("MSProject.Application", pythoncom.CoInitialize())
        _abs_path = os.path.abspath(path)
        print(_abs_path)
        msp.FileOpen(_abs_path)
        project = msp.ActiveProject
    except Exception:
        logging.error('%s: Файл проекта не смог открыться', get_project.__name__)
        raise Exception('Не получилось открыть файл проекта')
    logging.info('%s: Файл проекта успешно открылся', get_project.__name__)
    return project, msp


def get_excel_pd(path):
    """Создает DataFrame из ОФ и заменяет пустые строки на значение, указанное в config"""
    if not os.path.isabs(path):
        logging.warning('%s: Путь до ОФ не абсолютный', get_excel_pd.__name__)
    logging.info('%s: Пытаемся записать ОФ в DataFrame', get_excel_pd.__name__)
    try:
        data = pd.read_excel(path)
        data.fillna(value=config.errors.get('None'), inplace=True)
        for key, value in config.errors.items():
            if key != 'None':
                data.replace(key, value, inplace=True)
    except Exception:
        logging.error('%s: Не получилось записать ОФ в DataFrame', get_project.__name__)
        raise Exception('Не получилось записать ОФ в DataFrame')
    logging.info('%s: ОФ успешно записалась в DataFrame', get_excel_pd.__name__)
    return data


def get_project_pd(project, columns):
    """Создает DataFrame из столбцов объекта проекта."""
    logging.info('%s: Создаем DataFrame из столбцов объекта проекта', get_project_pd.__name__)
    if not project:
        logging.error('%s: Не удалось получить объект проекта', get_project.__name__)
        raise Exception("Объект проекта пустой")
    if not columns:
        logging.error('%s: Ключевые столбцы не заданы', get_project.__name__)
        raise Exception("Ключевые столбцы не заданы")
    task_collection = project.Tasks
    data = pd.DataFrame(columns=columns)
    data[columns[1]] = pd.to_datetime(data[columns[1]], dayfirst=True)
    data[columns[2]] = pd.to_datetime(data[columns[2]], dayfirst=True)
    try:
        for t in task_collection:
            if t.ActualStart != 'НД':
                actual_start = datetime.datetime.date(t.ActualStart)
            else:
                actual_start = 'НД'
            if t.ActualFinish != 'НД':
                actual_finish = datetime.datetime.date(t.ActualFinish)
            else:
                actual_finish = 'НД'
            data.loc[len(data.index)] = [t.Text4, actual_start, actual_finish]
    except Exception:
        logging.error('%s: Не получилось создать DataFrame из столбцов объекта проекта', get_project.__name__)
        raise Exception('Не получилось создать DataFrame из проекта')
    logging.info('%s: DataFrame из столбцов объекта проекта успешно создан', get_project_pd.__name__)
    return data


def change_project(project, msp, changes):
    """Вносит изменения в проект"""
    if not project:
        logging.info('%s: Объект проекта пустой', change_project.__name__)
        raise Exception('Объект проекта пустой')
    task_collection = project.Tasks
    if not changes:
        logging.info('%s: Изменений в проекте нет', change_project.__name__)
    else:
        logging.info('%s: Применяем изменения', change_project.__name__)
        try:
            for i, t in enumerate(task_collection):
                if i in changes.keys():
                    print(type(t.ActualStart))
                    print(type(changes[i][1]))
                    #Не получается присвоить значение в t.ActualStart.
                    #t.ActualStart = t.ActualStart.replace(day=changes[i][1])
                    #pywintypes.datetime и datetime.date
                    t.Name = t.Name.replace('п', '1')
            msp.FileSave()
        except Exception:
            logging.error('%s: Не получилось применить изменения', change_project.__name__)
            raise Exception('Не получилось применить изменения')
        logging.info('%s: Изменения успешно применены', change_project.__name__)



def check_str(excel_str, project_str, columns):
    """Сравнивает значения task ОФ и проекта"""
    for col in columns:
        if not pd.isnull(excel_str[col]) and excel_str[col] != 'НД':
            excel_str[col] = datetime.datetime.date(excel_str[col])
        if excel_str[col] != project_str[col]:
            return False, col
    return True, None


def paint_excel(ws, i, column):
    """Закрашивает ячейку в ОФ, которая не совпадает с проектом"""
    letter = chr(65 + column)
    work_sheet_a1 = ws[f"{letter}{i}"]
    work_sheet_a1.fill = opx.styles.PatternFill(fill_type='solid', start_color=config.color_fill)


def check_form(data_project, data_excel, columns):
    """Находит несоотвествия между ОФ и проектом и сохраняет их в словарь"""
    logging.info('%s: Ищем несоответсвия между ОФ и проектом', check_form.__name__)
    try:
        wb = opx.load_workbook(r"051-2000260_оф_ф_(13.04).xlsx")
        ws = wb.active
    except Exception:
        raise Exception('Не получилось открыть ОФ для заливки ячеек')
    if data_project.empty:
        logging.error('%s: DataFrame проекта пустой', check_form.__name__)
        raise Exception('DataFrame проекта пустой')
    if data_excel.empty:
        logging.error('%s: DataFrame ОФ пустой', check_form.__name__)
        raise Exception('DataFrame ОФ пустой')
    changes = {}
    for i, excel_str in data_excel.iterrows():
        for j, project_str in data_project.iterrows():
            if excel_str[columns[0]] == project_str[columns[0]]:
                status, column = check_str(excel_str, project_str, columns[1:])
                if not status:
                    paint_excel(ws, i + 2, data_excel.columns.get_loc(column))
                    changes[j] = [column, excel_str[column]]
                break
    wb.save(r"Изменения в ОФ.xlsx")
    logging.info('%s: Поиск несоответсвия между ОФ и проектом окончен', check_form.__name__)
    return changes


def main():

    logging.basicConfig(level=logging.DEBUG, stream=sys.stdout)

    path_to_project = '051-2000260_2022_нг_ф_(06.04).mpp'
    path_to_excel = '051-2000260_оф_ф_(13.04).xlsx'

    try:
        project, msp = get_project(path_to_project)
        data_project = get_project_pd(project, config.columns)
        data_excel = get_excel_pd(path_to_excel)
        changes = check_form(data_project, data_excel, config.columns)
        change_project(project, msp, changes)
    except Exception as e:
        print(e)

    logging.info('Работа программы закончена')

main()
